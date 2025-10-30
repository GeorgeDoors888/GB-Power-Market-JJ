#!/usr/bin/env python3
"""
Extract comprehensive tariff data from ALL 14 UK DNOs (2020-2026)
Includes: LLFCs, PCs, Red/Amber/Green rates, Fixed/Capacity charges, Document references
"""

import pandas as pd
import re
from pathlib import Path
from typing import List, Dict, Any
import openpyxl

# DNO Code Mappings
DNO_CODE_MAP = {
    # NGED (National Grid Electricity Distribution) - 4 DNOs
    'EMEB': {'code': 'EMID', 'name': 'East Midlands'},
    'MIDE': {'code': 'WMID', 'name': 'West Midlands'},
    'SWAE': {'code': 'SWAE', 'name': 'South Wales'},
    'SWEB': {'code': 'SWEB', 'name': 'South West'},
    
    # UKPN (UK Power Networks) - 3 DNOs
    'london': {'code': 'LPN', 'name': 'London'},
    'eastern': {'code': 'EPN', 'name': 'Eastern'},
    'south-eastern': {'code': 'SPN', 'name': 'South Eastern'},
    
    # Northern Powergrid - 2 DNOs
    'Northeast': {'code': 'NEPN', 'name': 'Northern Powergrid Northeast'},
    'Yorkshire': {'code': 'YPED', 'name': 'Northern Powergrid Yorkshire'},
    
    # ENWL (Electricity North West) - 1 DNO
    'enwl': {'code': 'ENWL', 'name': 'Electricity North West'},
    
    # SP Energy Networks - 2 DNOs
    'SPD': {'code': 'SPD', 'name': 'SP Distribution'},
    'SPM': {'code': 'SPM', 'name': 'SP Manweb'},
    
    # SSE (Scottish and Southern Energy) - 2 DNOs
    'shepd': {'code': 'SHEPD', 'name': 'Scottish Hydro'},
    'sepd': {'code': 'SEPD', 'name': 'Southern Electric'},
}


def extract_document_reference(filename: str) -> str:
    """Extract version and year from filename"""
    # Extract version (v0.1, V.0.2, v1.2, etc.)
    version_match = re.search(r'[Vv]\.?(\d+\.?\d*)', filename)
    version = f"v{version_match.group(1)}" if version_match else ""
    
    # Extract year (2020-2026)
    year_match = re.search(r'20(2[0-6])', filename)
    year = f"20{year_match.group(1)}" if year_match else ""
    
    return f"{year} {version}".strip()


def identify_dno_from_filename(filename: str) -> Dict[str, str]:
    """Identify DNO from filename"""
    filename_lower = filename.lower()
    
    for key, info in DNO_CODE_MAP.items():
        if key.lower() in filename_lower:
            return {'code': info['code'], 'name': info['name']}
    
    return {'code': 'UNKNOWN', 'name': 'Unknown DNO'}


def clean_value(val) -> str:
    """Clean cell value"""
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.lower() in ['nan', 'none', '']:
        return ""
    return val_str


def clean_rate(val) -> float:
    """Clean and convert rate value to float"""
    if pd.isna(val):
        return 0.0
    val_str = str(val).strip()
    if val_str.lower() in ['nan', 'none', '', 'n/a', '-']:
        return 0.0
    try:
        # Remove any currency symbols or commas
        val_str = val_str.replace('£', '').replace(',', '').replace('p', '').strip()
        return float(val_str)
    except:
        return 0.0


def find_tariff_sheet(workbook: openpyxl.Workbook) -> str:
    """Find the sheet with tariff data"""
    # Priority order for sheet names
    priority_sheets = [
        'Annex 1 LV, HV and UMS charges',
        'Annex 1',
        'LV charges',
        'Schedule of charges',
        'Tariff table',
        'Charges',
    ]
    
    sheet_names = [s.lower() for s in workbook.sheetnames]
    
    # Check priority sheets first
    for priority in priority_sheets:
        if priority.lower() in sheet_names:
            idx = sheet_names.index(priority.lower())
            return workbook.sheetnames[idx]
    
    # Look for sheets containing key terms
    for sheet_name in workbook.sheetnames:
        sheet_lower = sheet_name.lower()
        if any(term in sheet_lower for term in ['annex', 'charges', 'tariff', 'schedule']):
            return sheet_name
    
    # Default to first sheet
    return workbook.sheetnames[0] if workbook.sheetnames else None


def find_header_row(df: pd.DataFrame) -> int:
    """Find the header row containing tariff column names"""
    search_terms = ['tariff', 'llfc', 'profile class', 'pc', 'red', 'amber', 'green', 'unit rate']
    
    for idx, row in df.iterrows():
        row_str = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
        if sum(term in row_str for term in search_terms) >= 3:
            return idx
    
    return 0  # Default to first row


def extract_tariffs_from_file(file_path: Path) -> List[Dict[str, Any]]:
    """Extract tariffs from a single Excel file"""
    print(f"   📂 {file_path.name}")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = find_tariff_sheet(wb)
        
        if not sheet_name:
            print(f"      ⚠️  No suitable sheet found")
            return []
        
        # Read sheet into DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Find header row
        header_row = find_header_row(df)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        
        # Identify DNO from filename
        dno_info = identify_dno_from_filename(file_path.name)
        
        # Extract document reference
        doc_ref = extract_document_reference(file_path.name)
        
        # Extract year from filename or doc_ref
        year_match = re.search(r'20(2[0-6])', file_path.name)
        year = f"20{year_match.group(1)}" if year_match else doc_ref.split()[0] if doc_ref else ""
        
        tariffs = []
        
        # Clean column names (remove line breaks, extra spaces)
        df.columns = [str(col).replace('\n', ' ').replace('\r', ' ').strip() for col in df.columns]
        
        # Column name variations to look for
        tariff_cols = ['tariff', 'tariff name', 'tariff description', 'description']
        llfc_cols = ['llfc', 'llfcs', 'line loss factor class', 'llf class', 'open llfc']
        pc_cols = ['pc', 'pcs', 'profile class', 'profile classes']
        red_cols = ['red', 'red rate', 'red unit rate', 'red (p/kwh)', 'red unit charge', 'super red', 'peak', 'black unit charge']
        amber_cols = ['amber', 'amber rate', 'amber unit rate', 'amber (p/kwh)', 'amber unit charge', 'yellow unit charge']
        green_cols = ['green', 'green rate', 'green unit rate', 'green (p/kwh)', 'green unit charge', 'off peak', 'night']
        fixed_cols = ['fixed', 'fixed charge', 'standing charge', 'fixed (p/day)', 'fixed charge p/mpan/day', 'capacity']
        capacity_cols = ['capacity charge', 'capacity (p/kva/day)', 'capacity charge p/kva/day', 'exceeded capacity']
        
        # Find actual column names (case-insensitive, substring matching)
        def find_col(variations, df_cols):
            df_cols_lower = {col.lower(): col for col in df_cols}
            # First try exact match
            for var in variations:
                if var.lower() in df_cols_lower:
                    return df_cols_lower[var.lower()]
            # Then try substring match
            for var in variations:
                for col_lower, col_actual in df_cols_lower.items():
                    if var.lower() in col_lower:
                        return col_actual
            return None
        
        df_cols = df.columns.tolist()
        tariff_col = find_col(tariff_cols, df_cols)
        llfc_col = find_col(llfc_cols, df_cols)
        pc_col = find_col(pc_cols, df_cols)
        red_col = find_col(red_cols, df_cols)
        amber_col = find_col(amber_cols, df_cols)
        green_col = find_col(green_cols, df_cols)
        fixed_col = find_col(fixed_cols, df_cols)
        capacity_col = find_col(capacity_cols, df_cols)
        
        #  Debug: print what columns were found
        print(f"      Columns found: tariff={tariff_col}, llfc={llfc_col}, red={red_col}, amber={amber_col}, green={green_col}")
        
        # Extract tariffs
        for _, row in df.iterrows():
            tariff_name = clean_value(row.get(tariff_col, "")) if tariff_col else ""
            
            # Skip empty or header rows
            if not tariff_name or tariff_name.lower() in ['tariff', 'description', 'tariff name']:
                continue
            
            # Skip if no rate data
            red_rate = clean_rate(row.get(red_col, 0)) if red_col else 0.0
            amber_rate = clean_rate(row.get(amber_col, 0)) if amber_col else 0.0
            green_rate = clean_rate(row.get(green_col, 0)) if green_col else 0.0
            
            if red_rate == 0 and amber_rate == 0 and green_rate == 0:
                continue
            
            tariff = {
                'Year': year,
                'DNO_Code': dno_info['code'],
                'DNO_Name': dno_info['name'],
                'Tariff_Name': tariff_name,
                'LLFCs': clean_value(row.get(llfc_col, "")) if llfc_col else "",
                'PCs': clean_value(row.get(pc_col, "")) if pc_col else "",
                'Red_Rate_p_kWh': red_rate,
                'Amber_Rate_p_kWh': amber_rate,
                'Green_Rate_p_kWh': green_rate,
                'Fixed_Charge_p_day': clean_rate(row.get(fixed_col, 0)) if fixed_col else 0.0,
                'Capacity_Charge_p_kVA_day': clean_rate(row.get(capacity_col, 0)) if capacity_col else 0.0,
                'Document': file_path.name,
                'Document_Reference': doc_ref,
            }
            
            tariffs.append(tariff)
        
        print(f"      ✅ Extracted {len(tariffs)} tariffs")
        return tariffs
        
    except Exception as e:
        print(f"      ❌ Error: {e}")
        return []


def main():
    """Main extraction function"""
    print("\n" + "="*80)
    print("📊 COMPREHENSIVE DNO TARIFF EXTRACTION - ALL 14 DNOs")
    print("="*80 + "\n")
    
    base_path = Path(".")
    all_tariffs = []
    
    # Define file locations for each DNO
    file_locations = {
        'NGED': 'duos_nged_data/',
        'UKPN': '.',  # Root directory
        'Northern_Powergrid': '.',  # Root directory
        'ENWL': '.',  # Root directory
        'SPEN': 'duos_spm_data/',
        'SSE': 'duos_ssen_data/',
    }
    
    # 1. NGED (4 DNOs)
    print("🔍 NGED (National Grid Electricity Distribution)")
    nged_path = base_path / file_locations['NGED']
    if nged_path.exists():
        nged_files = sorted(nged_path.glob('*.xlsx'))
        for file in nged_files:
            if not file.name.startswith('~'):
                tariffs = extract_tariffs_from_file(file)
                all_tariffs.extend(tariffs)
    
    # 2. UKPN (3 DNOs)
    print("\n🔍 UK Power Networks (London, Eastern, South Eastern)")
    for pattern in ['london-power-networks-*.xlsx', 'eastern-power-networks-*.xlsx', 'south-eastern-power-networks-*.xlsx']:
        ukpn_files = sorted(base_path.glob(pattern))
        for file in ukpn_files:
            if not file.name.startswith('~'):
                tariffs = extract_tariffs_from_file(file)
                all_tariffs.extend(tariffs)
    
    # 3. Northern Powergrid (2 DNOs)
    print("\n🔍 Northern Powergrid (Northeast, Yorkshire)")
    npg_files = sorted(base_path.glob('Northern Powergrid*.xlsx'))
    for file in npg_files:
        if not file.name.startswith('~') and 'Schedule of charges' in file.name:
            tariffs = extract_tariffs_from_file(file)
            all_tariffs.extend(tariffs)
    
    # 4. ENWL (1 DNO)
    print("\n🔍 Electricity North West")
    enwl_files = sorted(base_path.glob('enwl*.xlsx'))
    for file in enwl_files:
        if not file.name.startswith('~') and 'schedule' in file.name.lower():
            tariffs = extract_tariffs_from_file(file)
            all_tariffs.extend(tariffs)
    
    # 5. SP Energy Networks (2 DNOs)
    print("\n🔍 SP Energy Networks (SPD, SPM)")
    spen_path = base_path / file_locations['SPEN']
    if spen_path.exists():
        spen_files = sorted(spen_path.glob('SP*.xlsx'))
        for file in spen_files:
            if not file.name.startswith('~') and 'Schedule' in file.name:
                tariffs = extract_tariffs_from_file(file)
                all_tariffs.extend(tariffs)
    
    # Also check root for SPEN files
    for pattern in ['SPD*.xlsx', 'SPM*.xlsx', 'SP_Distribution*.xlsx', 'SP_Manweb*.xlsx']:
        spen_files = sorted(base_path.glob(pattern))
        for file in spen_files:
            if not file.name.startswith('~'):
                tariffs = extract_tariffs_from_file(file)
                all_tariffs.extend(tariffs)
    
    # 6. SSE (2 DNOs)
    print("\n🔍 SSE (Scottish Hydro, Southern Electric)")
    sse_path = base_path / file_locations['SSE']
    if sse_path.exists():
        for subdir in ['shepd', 'sepd']:
            sse_subpath = sse_path / subdir
            if sse_subpath.exists():
                sse_files = sorted(sse_subpath.glob('*.xlsx'))
                for file in sse_files:
                    if not file.name.startswith('~') and 'schedule' in file.name.lower():
                        tariffs = extract_tariffs_from_file(file)
                        all_tariffs.extend(tariffs)
    
    # Create DataFrame
    df = pd.DataFrame(all_tariffs)
    
    # Summary
    print("\n" + "="*80)
    print("📊 EXTRACTION SUMMARY")
    print("="*80)
    print(f"Total tariffs extracted: {len(df)}")
    
    if len(df) > 0:
        print(f"\nDNOs covered: {df['DNO_Name'].nunique()}")
        print(df.groupby('DNO_Name')['Year'].apply(lambda x: sorted(x.unique())).to_string())
        
        print("\n📊 Tariff counts by DNO and Year:")
        summary = df.groupby(['DNO_Name', 'Year']).size().reset_index(name='Count')
        print(summary.to_string(index=False))
        
        # Save to CSV
        output_csv = 'all_14_dnos_comprehensive_tariffs.csv'
        df.to_csv(output_csv, index=False)
        print(f"\n✅ Saved to: {output_csv}")
        
        # Save to Excel with sheets by year
        output_xlsx = 'all_14_dnos_comprehensive_tariffs.xlsx'
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            # All tariffs
            df.to_excel(writer, sheet_name='All Tariffs', index=False)
            
            # By year
            for year in sorted(df['Year'].unique()):
                if year:
                    year_df = df[df['Year'] == year]
                    sheet_name = f'Year {year}'
                    year_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # By DNO
            for dno in sorted(df['DNO_Name'].unique()):
                dno_df = df[df['DNO_Name'] == dno]
                sheet_name = dno[:31]  # Excel sheet name limit
                dno_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"✅ Saved to: {output_xlsx}")
        
        # Show sample of traffic light tariffs
        print("\n📋 SAMPLE: Non-Domestic Traffic Light Tariffs")
        sample = df[df['Tariff_Name'].str.contains('Non-Domestic', case=False, na=False)]
        if len(sample) > 0:
            print(f"\nFound {len(sample)} Non-Domestic tariffs")
            sample_display = sample[['DNO_Name', 'Year', 'Tariff_Name', 'Red_Rate_p_kWh', 'Amber_Rate_p_kWh', 'Green_Rate_p_kWh']].head(10)
            print(sample_display.to_string(index=False))
    
    print("\n" + "="*80)
    print("✅ EXTRACTION COMPLETE!")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
