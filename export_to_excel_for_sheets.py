#!/usr/bin/env python3
"""
Export NGED Charging Data to Excel for Google Sheets Upload
Creates Excel files split by 10MB if needed, ready for manual upload to Google Drive
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
CSV_PATH = Path("/Users/georgemajor/GB Power Market JJ/nged_charging_data_parsed.csv")
OUTPUT_DIR = Path("/Users/georgemajor/GB Power Market JJ/google_sheets_ready")
MAX_FILE_SIZE_MB = 9  # Keep under 10MB for Google Sheets
BASE_FILENAME = "NGED_Charging_Data"

def load_data():
    """Load parsed CSV data"""
    print(f"📂 Loading data from CSV...")
    df = pd.read_csv(CSV_PATH)
    
    # Clean up the data for Excel
    if 'rates_found' in df.columns:
        df['rates_found'] = df['rates_found'].apply(lambda x: str(x) if pd.notna(x) else '')
    if 'values_found' in df.columns:
        df['values_found'] = df['values_found'].apply(lambda x: str(x) if pd.notna(x) else '')
    
    # Shorten long text fields
    if 'raw_text' in df.columns:
        df['raw_text'] = df['raw_text'].apply(lambda x: str(x)[:200] if pd.notna(x) else '')
    
    print(f"   ✅ Loaded {len(df):,} records")
    print(f"   📊 Columns: {len(df.columns)}")
    return df

def create_summary_data(df):
    """Create summary statistics"""
    summary_data = {
        'Metric': [],
        'Value': []
    }
    
    summary_data['Metric'].extend([
        'Total Records',
        'Years Covered',
        'DNO Areas',
        'Files Parsed',
        'Sheets Parsed',
        '',
        'Records by Year'
    ])
    
    summary_data['Value'].extend([
        len(df),
        f"{df['year'].min()} - {df['year'].max()}",
        ', '.join(df['dno_code'].unique()),
        df['filename'].nunique(),
        df['sheet'].nunique(),
        '',
        ''
    ])
    
    # Year breakdown
    year_counts = df['year'].value_counts().sort_index()
    for year, count in year_counts.items():
        summary_data['Metric'].append(f'  {year}')
        summary_data['Value'].append(count)
    
    summary_data['Metric'].append('')
    summary_data['Value'].append('')
    summary_data['Metric'].append('Records by DNO')
    summary_data['Value'].append('')
    
    # DNO breakdown
    dno_counts = df.groupby(['dno_code', 'dno_name']).size().sort_values(ascending=False)
    for (dno, name), count in dno_counts.items():
        summary_data['Metric'].append(f'  {dno} ({name})')
        summary_data['Value'].append(count)
    
    return pd.DataFrame(summary_data)

def format_worksheet(ws, df, is_summary=False):
    """Apply formatting to worksheet"""
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max 50 chars
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply filters
    ws.auto_filter.ref = ws.dimensions

def create_excel_file(df, output_path, part=None):
    """Create Excel file with data and summary"""
    print(f"   📝 Creating Excel file: {output_path.name}")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet
    ws_summary = wb.create_sheet('Summary', 0)
    summary_df = create_summary_data(df)
    
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)
    
    format_worksheet(ws_summary, summary_df, is_summary=True)
    
    # Add title
    ws_summary.insert_rows(0)
    ws_summary['A1'] = 'NGED Charging Data Summary'
    ws_summary['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_summary['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Create Data sheet
    ws_data = wb.create_sheet('Charging Data', 1)
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    format_worksheet(ws_data, df)
    
    # Save
    wb.save(output_path)
    
    file_size_mb = output_path.stat().st_size / (1024 * 1024)
    print(f"      ✅ Saved: {file_size_mb:.2f} MB")
    
    return file_size_mb

def split_data_by_size(df):
    """Split data into chunks that fit size limit"""
    # Estimate size per row (rough estimate)
    test_file = OUTPUT_DIR / "test.xlsx"
    test_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Test with small sample
    sample_df = df.head(1000)
    create_excel_file(sample_df, test_file)
    
    sample_size_mb = test_file.stat().st_size / (1024 * 1024)
    bytes_per_row = (sample_size_mb * 1024 * 1024) / 1000
    
    test_file.unlink()
    
    # Calculate rows per file
    max_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
    rows_per_file = int(max_bytes / bytes_per_row * 0.8)  # 80% safety margin
    
    print(f"   📊 Estimated {bytes_per_row:.0f} bytes per row")
    print(f"   📊 Target: ~{rows_per_file:,} rows per file")
    
    # Split data
    splits = []
    total_rows = len(df)
    
    if rows_per_file >= total_rows:
        splits.append({'df': df, 'name': f"{BASE_FILENAME}.xlsx", 'part': None})
    else:
        num_parts = (total_rows + rows_per_file - 1) // rows_per_file
        for i in range(num_parts):
            start_idx = i * rows_per_file
            end_idx = min((i + 1) * rows_per_file, total_rows)
            chunk = df.iloc[start_idx:end_idx].copy()
            
            splits.append({
                'df': chunk,
                'name': f"{BASE_FILENAME}_Part_{i+1}_of_{num_parts}.xlsx",
                'part': i + 1
            })
    
    return splits

def main():
    """Main execution"""
    print("=" * 80)
    print("NGED CHARGING DATA - EXCEL EXPORT FOR GOOGLE SHEETS")
    print("=" * 80)
    print()
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"📁 Output directory: {OUTPUT_DIR}")
    print()
    
    # Load data
    df = load_data()
    print()
    
    # Split data by size
    print(f"📦 Calculating file splits...")
    splits = split_data_by_size(df)
    print(f"   ✅ Will create {len(splits)} file(s)")
    print()
    
    # Create files
    created_files = []
    
    for i, split_info in enumerate(splits, 1):
        print(f"[{i}/{len(splits)}] Processing: {split_info['name']}")
        output_path = OUTPUT_DIR / split_info['name']
        
        file_size = create_excel_file(split_info['df'], output_path, split_info['part'])
        
        created_files.append({
            'name': split_info['name'],
            'path': output_path,
            'size_mb': file_size,
            'rows': len(split_info['df'])
        })
        print()
    
    # Final summary
    print("=" * 80)
    print("✅ EXCEL FILES CREATED!")
    print("=" * 80)
    print()
    print(f"📊 Summary:")
    print(f"   Total records: {len(df):,}")
    print(f"   Files created: {len(created_files)}")
    print()
    
    print("📄 Files ready for upload:")
    for file_info in created_files:
        print(f"   • {file_info['name']}")
        print(f"     Size: {file_info['size_mb']:.2f} MB | Rows: {file_info['rows']:,}")
        print(f"     Path: {file_info['path']}")
        print()
    
    print("=" * 80)
    print("📤 UPLOAD INSTRUCTIONS")
    print("=" * 80)
    print()
    print("1. Open Google Drive in your browser:")
    print("   https://drive.google.com")
    print()
    print("2. Create a new folder (optional): 'DNO Charging Data'")
    print()
    print("3. Upload the Excel file(s):")
    print(f"   Drag and drop from: {OUTPUT_DIR}")
    print()
    print("4. Once uploaded, right-click each file and select:")
    print("   'Open with' → 'Google Sheets'")
    print()
    print("5. Google Sheets will convert the Excel file automatically")
    print()
    print("6. The Summary sheet will show statistics")
    print("   The Charging Data sheet contains all records")
    print()
    print("💡 Benefits of this approach:")
    print("   ✓ Uses YOUR Google Drive storage (7TB available)")
    print("   ✓ No service account quota issues")
    print("   ✓ Full control over sharing and permissions")
    print("   ✓ Can use Google Sheets features (filters, pivots, formulas)")
    print()

if __name__ == "__main__":
    main()
