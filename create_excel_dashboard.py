#!/usr/bin/env python3
"""
Create Excel Dashboard from Google Sheets
Recreates GB Power Market dashboard with Excel-native formatting and functions
"""

import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.marker import DataPoint
import datetime
import re

# Google Sheets credentials
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SHEET_ID = '1-u794iGngn5_Ql_XocKSwvHSKWABWO0bVsudkUJAFqA'

# Excel color scheme (matching Google Sheets theme)
COLORS = {
    'primary_blue': '1F4E78',
    'accent_blue': '4472C4',
    'light_blue': 'D9E1F2',
    'orange': 'FF6600',
    'green': '70AD47',
    'red': 'FF0000',
    'yellow': 'FFC000',
    'gray': 'E7E6E6',
    'dark_gray': '808080',
    'white': 'FFFFFF'
}

class ExcelDashboardCreator:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        
        # Connect to Google Sheets
        creds = Credentials.from_service_account_file(
            'inner-cinema-credentials.json',
            scopes=SCOPES
        )
        self.gc = gspread.authorize(creds)
        self.spreadsheet = self.gc.open_by_key(SHEET_ID)
        
    def create_dashboard(self):
        """Create the main dashboard sheet with KPIs and formatting"""
        print("\n📊 Creating Live Dashboard v2...")
        
        gs_ws = self.spreadsheet.worksheet('Live Dashboard v2')
        data = gs_ws.get_all_values()
        
        ws = self.wb.create_sheet(title='Live Dashboard v2', index=0)
        
        # Write all data first
        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if value:
                    # Try numeric conversion
                    try:
                        cell.value = float(value)
                    except:
                        cell.value = value
                        
        # Format main header (Row 1)
        ws.merge_cells('A1:L1')
        cell = ws['A1']
        cell.font = Font(name='Calibri', size=20, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['orange'], end_color=COLORS['orange'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Format timestamp (Row 2)
        ws.merge_cells('A2:L2')
        cell = ws['A2']
        cell.font = Font(name='Calibri', size=10, italic=True, color=COLORS['dark_gray'])
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format section headers
        section_headers = [4, 10, 30]  # Row numbers with section headers
        for row_num in section_headers:
            cell = ws.cell(row=row_num, column=1)
            if cell.value:
                ws.merge_cells(f'A{row_num}:L{row_num}')
                cell.font = Font(name='Calibri', size=14, bold=True, color=COLORS['white'])
                cell.fill = PatternFill(start_color=COLORS['primary_blue'], end_color=COLORS['primary_blue'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                ws.row_dimensions[row_num].height = 25
        
        # Format KPI cards (Row 5 - assuming KPI row)
        for col in range(1, 13, 2):  # Every other column for KPI labels
            cell = ws.cell(row=5, column=col)
            if cell.value:
                cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['dark_gray'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format corresponding value cell
                value_cell = ws.cell(row=6, column=col)
                value_cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['accent_blue'])
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column[:100]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        print(f"   ✅ Dashboard created with {len(data)} rows")
        return ws
    
    def create_btm_calculator(self):
        """Create BtM Calculator with Excel formulas"""
        print("\n🧮 Creating BtM Calculator...")
        
        gs_ws = self.spreadsheet.worksheet('BtM Calculator')
        data = gs_ws.get_all_values()
        
        ws = self.wb.create_sheet(title='BtM Calculator')
        
        # Write data
        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if value:
                    try:
                        cell.value = float(value)
                    except:
                        cell.value = value
        
        # Format header
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['primary_blue'], end_color=COLORS['primary_blue'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Format input sections with light yellow background
        for row in range(4, 20):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.value and 'Input' in str(cell.value):
                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    cell.font = Font(bold=True)
        
        # Format output sections with light green background
        for row in range(20, 40):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.value and ('Output' in str(cell.value) or 'Revenue' in str(cell.value)):
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    cell.font = Font(bold=True)
        
        print(f"   ✅ Calculator created with {len(data)} rows")
        return ws
    
    def create_revenue_analysis(self):
        """Create BM Revenue Analysis with charts"""
        print("\n💰 Creating BM Revenue Analysis...")
        
        gs_ws = self.spreadsheet.worksheet('BM Revenue Analysis')
        data = gs_ws.get_all_values()
        
        ws = self.wb.create_sheet(title='BM Revenue Analysis')
        
        # Write data
        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if value:
                    try:
                        # Check if it's a currency value
                        if '£' in str(value) or '$' in str(value):
                            numeric_val = re.sub(r'[£$,]', '', str(value))
                            try:
                                cell.value = float(numeric_val)
                                cell.number_format = '"£"#,##0.00'
                            except:
                                cell.value = value
                        else:
                            cell.value = float(value)
                    except:
                        cell.value = value
        
        # Format header
        ws.merge_cells('A1:S1')
        cell = ws['A1']
        cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['orange'], end_color=COLORS['orange'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Format column headers (Row 2)
        for col in range(1, 20):
            cell = ws.cell(row=2, column=col)
            if cell.value:
                cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['white'])
                cell.fill = PatternFill(start_color=COLORS['accent_blue'], end_color=COLORS['accent_blue'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add conditional formatting for revenue values
        # (Note: openpyxl supports conditional formatting via ws.conditional_formatting.add())
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column[:50]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        print(f"   ✅ Revenue analysis created with {len(data)} rows")
        return ws
    
    def create_boalf_analysis(self):
        """Create BOALF Price Analysis"""
        print("\n📈 Creating BOALF Price Analysis...")
        
        gs_ws = self.spreadsheet.worksheet('BOALF Price Analysis')
        data = gs_ws.get_all_values()
        
        ws = self.wb.create_sheet(title='BOALF Price Analysis')
        
        # Write data
        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if value:
                    try:
                        cell.value = float(value)
                    except:
                        cell.value = value
        
        # Format headers
        for col in range(1, 40):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['white'])
                cell.fill = PatternFill(start_color=COLORS['primary_blue'], end_color=COLORS['primary_blue'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        print(f"   ✅ BOALF analysis created with {len(data)} rows")
        return ws
    
    def create_definitions_sheet(self):
        """Create BID/OFFER Definitions reference"""
        print("\n📖 Creating BID/OFFER Definitions...")
        
        gs_ws = self.spreadsheet.worksheet('BID_OFFER_Definitions')
        data = gs_ws.get_all_values()
        
        ws = self.wb.create_sheet(title='BID_OFFER_Definitions')
        
        # Write data
        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                
                # Format definition headers
                if r_idx == 1 or (cell.value and ('⚡' in str(cell.value) or '📖' in str(cell.value))):
                    cell.font = Font(name='Calibri', size=14, bold=True, color=COLORS['white'])
                    cell.fill = PatternFill(start_color=COLORS['orange'], end_color=COLORS['orange'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column[:100]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 15)
        
        print(f"   ✅ Definitions created with {len(data)} rows")
        return ws
    
    def create_scrp_summary(self):
        """Create SCRP Summary sheet"""
        print("\n⚡ Creating SCRP Summary...")
        
        gs_ws = self.spreadsheet.worksheet('SCRP_Summary')
        data = gs_ws.get_all_values()
        
        ws = self.wb.create_sheet(title='SCRP_Summary')
        
        # Write data
        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if value:
                    try:
                        cell.value = float(value)
                    except:
                        cell.value = value
        
        # Format main header
        if data and data[0]:
            ws.merge_cells(f'A1:B1')
            cell = ws['A1']
            cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['white'])
            cell.fill = PatternFill(start_color=COLORS['primary_blue'], end_color=COLORS['primary_blue'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
        
        print(f"   ✅ SCRP Summary created with {len(data)} rows")
        return ws
    
    def save(self, filename='GB_Power_Market_Dashboard_Enhanced.xlsx'):
        """Save the workbook"""
        self.wb.save(filename)
        print(f"\n{'=' * 70}")
        print(f"✅ SUCCESS: Saved {filename}")
        print(f"📊 Total sheets: {len(self.wb.sheetnames)}")
        for sheet in self.wb.sheetnames:
            print(f"   • {sheet}")
        print(f"\n📁 Location: /home/george/GB-Power-Market-JJ/{filename}")
        return filename

def main():
    print("🚀 GB Power Market Dashboard - Excel Recreation")
    print("=" * 70)
    
    creator = ExcelDashboardCreator()
    
    # Create all sheets
    creator.create_dashboard()
    creator.create_btm_calculator()
    creator.create_revenue_analysis()
    creator.create_boalf_analysis()
    creator.create_definitions_sheet()
    creator.create_scrp_summary()
    
    # Save workbook
    output_file = creator.save()
    
    print("\n✨ Excel dashboard creation complete!")
    print(f"\nTo open: libreoffice {output_file}")

if __name__ == '__main__':
    main()
