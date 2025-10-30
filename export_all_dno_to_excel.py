#!/usr/bin/env python3
"""
Export ALL DNO Charging Data to Excel for Google Sheets Upload
Handles auto-splitting if files exceed 10MB
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configuration
CSV_PATH = Path("/Users/georgemajor/GB Power Market JJ/all_dno_charging_data_parsed.csv")
OUTPUT_DIR = Path("/Users/georgemajor/GB Power Market JJ/google_sheets_ready")
MAX_FILE_SIZE_MB = 9  # Safety margin below 10MB limit

def load_data():
    """Load the parsed CSV data"""
    print("📂 Loading ALL DNO charging data...")
    df = pd.read_csv(CSV_PATH)
    
    # Clean up NaN values for Google Sheets
    df = df.fillna('')
    
    print(f"   ✅ Loaded {len(df):,} records")
    print(f"   📊 DNOs: {', '.join(df['dno_code'].unique())}")
    print(f"   📅 Years: {df['year'].min()} - {df['year'].max()}")
    
    return df

def split_data_by_size(df):
    """Split data into chunks that stay under size limit"""
    print()
    print(f"📏 Calculating file sizes...")
    
    # Estimate bytes per row from sample
    sample_size = min(1000, len(df))
    sample_df = df.head(sample_size)
    
    # Save sample to estimate size
    sample_path = OUTPUT_DIR / "temp_sample.xlsx"
    with pd.ExcelWriter(sample_path, engine='openpyxl') as writer:
        sample_df.to_excel(writer, index=False, sheet_name='Data')
    
    sample_size_bytes = sample_path.stat().st_size
    bytes_per_row = sample_size_bytes / sample_size
    sample_path.unlink()  # Delete temp file
    
    # Calculate rows per file
    max_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
    rows_per_file = int(max_bytes / bytes_per_row)
    
    print(f"   📊 Estimated {bytes_per_row:.0f} bytes per row")
    print(f"   📦 Will split into chunks of {rows_per_file:,} rows")
    
    # Split data
    splits = []
    total_rows = len(df)
    num_files = (total_rows + rows_per_file - 1) // rows_per_file
    
    for i in range(num_files):
        start_idx = i * rows_per_file
        end_idx = min((i + 1) * rows_per_file, total_rows)
        chunk = df.iloc[start_idx:end_idx]
        
        splits.append({
            'name': f"All_DNO_Charging_Data_Part_{i+1}_of_{num_files}" if num_files > 1 else "All_DNO_Charging_Data",
            'data': chunk,
            'part': i + 1,
            'total_parts': num_files
        })
    
    print(f"   ✅ Data will be split into {num_files} file(s)")
    print()
    
    return splits

def create_excel_file(df, output_path, part_info=None):
    """Create formatted Excel file"""
    print(f"   📄 Creating Excel file: {output_path.name}")
    
    # Create workbook
    wb = Workbook()
    
    # Create Summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Write summary info
    summary_data = [
        ["All DNO Charging Data Summary", ""],
        ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
    ]
    
    if part_info:
        summary_data.extend([
            ["File Part", f"{part_info['part']} of {part_info['total_parts']}"],
            ["Records in this file", len(df)],
            ["", ""],
        ])
    
    summary_data.extend([
        ["Statistics", ""],
        ["Total Records", len(df)],
        ["Years Covered", f"{df['year'].min()} - {df['year'].max()}"],
        ["DNO Areas", len(df['dno_code'].unique())],
        ["", ""],
        ["Records by DNO", ""],
    ])
    
    # Add DNO breakdown
    dno_counts = df['dno_code'].value_counts().sort_index()
    for dno, count in dno_counts.items():
        dno_name = df[df['dno_code'] == dno]['dno_name'].iloc[0]
        summary_data.append([f"  {dno} ({dno_name})", int(count)])
    
    summary_data.append(["", ""])
    summary_data.append(["Records by Year", ""])
    
    # Add year breakdown
    year_counts = df['year'].value_counts().sort_index()
    for year, count in year_counts.items():
        summary_data.append([f"  {year}", int(count)])
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Data sheet
    data_sheet = wb.create_sheet("Charging Data")
    
    # Write headers
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            data_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns (max 50 chars)
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row in data_sheet.iter_rows(min_row=2, max_row=min(100, len(df) + 1), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        data_sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Freeze header row
    data_sheet.freeze_panes = "A2"
    
    # Enable auto-filter
    data_sheet.auto_filter.ref = data_sheet.dimensions
    
    # Save
    wb.save(output_path)
    
    file_size_mb = output_path.stat().st_size / (1024 * 1024)
    print(f"   ✅ File created: {file_size_mb:.2f} MB")
    
    return file_size_mb

def main():
    """Main execution"""
    print("=" * 80)
    print("ALL DNO CHARGING DATA - EXCEL EXPORT")
    print("=" * 80)
    print()
    
    # Create output directory
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    # Load data
    df = load_data()
    
    # Split if needed
    splits = split_data_by_size(df)
    
    print(f"📦 Creating {len(splits)} Excel file(s)...")
    print()
    
    total_size = 0
    created_files = []
    
    for split in splits:
        output_path = OUTPUT_DIR / f"{split['name']}.xlsx"
        
        part_info = {
            'part': split['part'],
            'total_parts': split['total_parts']
        } if split['total_parts'] > 1 else None
        
        file_size = create_excel_file(split['data'], output_path, part_info)
        total_size += file_size
        created_files.append(output_path.name)
    
    print()
    print("=" * 80)
    print("✅ EXPORT COMPLETE!")
    print("=" * 80)
    print()
    print(f"📁 Output directory: {OUTPUT_DIR}")
    print(f"📦 Files created: {len(created_files)}")
    print()
    
    for filename in created_files:
        print(f"   📄 {filename}")
    
    print()
    print(f"💾 Total size: {total_size:.2f} MB")
    print()
    print("🚀 Next steps:")
    print("   1. Run: open '{}'".format(OUTPUT_DIR))
    print("   2. Drag files to drive.google.com (george@upowerenergy.uk)")
    print("   3. Right-click → 'Open with' → 'Google Sheets'")
    print()
    print("   OR use OAuth upload:")
    print("   .venv/bin/python upload_all_dno_with_oauth.py")
    print()

if __name__ == "__main__":
    main()
